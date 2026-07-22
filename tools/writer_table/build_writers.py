#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
전교인 성경필사 색인 생성기  ─  cropped/ (책별 폴더) → writers.json

파이프라인
  1) cropped/ 재귀 탐색 (하위 폴더명 = 책)
  2) 각 crop 이미지의 이름 텍스트 확보  (OCR)
       우선순위: ① 이미지 옆 사이드카 .txt  ② ocr_image() 훅  ③ 빈칸
  3) 분배표(Name/Book/FROM/TO)의 '정답 이름'에 자모 단위 퍼지 매칭 → 오류 보정
       └ 파일 경로의 '책'으로 후보를 한정해 동명이인·다중배정을 갈라냄
  4) 간혹 비어있는 이름칸은 앞뒤 페이지 맥락으로 채움 (경계는 확인필요로 표시)
  5) 같은 사람의 연속 페이지를 한 entry로 묶어 writer 단위 집계
  6) 검수용 필드(점수·원본OCR·needs_review 사유)와 함께 writers.json 출력

OCR 연동
  손글씨 OCR은 David 님이 이미 쓰는 엔진이 가장 정확합니다. 두 방법 중 택1:
    A) 각 이미지 옆에 같은 이름의 .txt 로 OCR 결과를 놔두면 그걸 그대로 읽음
       (예: amos/001.png  ↔  amos/001.txt 안에 "임종만")
    B) 아래 ocr_image() 함수 본문을 본인 엔진 호출로 교체
"""
import argparse, json, re, sys, unicodedata
from pathlib import Path
import pandas as pd
from rapidfuzz.distance import Levenshtein

# ══════════════════════════════════════════════════════════════════
# 0. 설정
# ══════════════════════════════════════════════════════════════════
IMG_EXT = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp"}
SCORE_NOTE = 80.0       # 이 점수 미만 → 단순 안내(note). 손글씨는 점수 낮아도 매칭은 맞는 경우 많음
SCORE_HARDFLOOR = 45.0  # 이 점수 미만 → 검수(OCR이 너무 깨져 신뢰 불가)
MARGIN_REVIEW = 10.0    # 1·2위 격차가 이보다 작으면 검수(후보끼리 애매 = 진짜 위험)
POS_CASEB_FLOOR = 30.0  # 경계 페이지를 위치로 자동 재배정하는 점수 상한(이보다 낮을 때만)

# 폴더명(영문) → 분배표의 한글 책명.  book_id 는 폴더명(영문)을 그대로 사용.
BOOK_KO = {
    "genesis":"창세기","exodus":"출애굽기","leviticus":"레위기","numbers":"민수기",
    "deuteronomy":"신명기","joshua":"여호수아","judges":"사사기","ruth":"룻기",
    "1samuel":"사무엘상","2samuel":"사무엘하","1kings":"열왕기상","2kings":"열왕기하",
    "1chronicles":"역대상","2chronicles":"역대하","ezra":"에스라","nehemiah":"느헤미야",
    "esther":"에스더","job":"욥기","psalms":"시 편","proverbs":"잠 언",
    "ecclesiastes":"전도서","songofsongs":"아가","isaiah":"이사야","jeremiah":"예레미야",
    # OCR 파일이 실제로 쓰는 서수-하이픈 표기도 함께 매핑
    "first-samuel":"사무엘상","second-samuel":"사무엘하",
    "first-kings":"열왕기상","second-kings":"열왕기하",
    "first-chronicles":"역대상","second-chronicles":"역대하",
    "song-of-songs":"아가",
    "lamentations":"예레미야애가","ezekiel":"에스겔","daniel":"다니엘","hosea":"호세아",
    "joel":"요엘","amos":"아모스","obadiah":"오바댜","jonah":"요나","micah":"미가",
    "nahum":"나훔","habakkuk":"하박국","zephaniah":"스바냐","haggai":"학개",
    "zechariah":"스가랴","malachi":"말라기",
    # 신약 (data/toc.json 의 new-testament book_id 순서/표기 기준)
    "matthew":"마태복음","mark":"마가복음","luke":"누가복음","john":"요한복음",
    "acts":"사도행전","romans":"로마서",
    "first-corinthians":"고린도전서","second-corinthians":"고린도후서",
    "galatians":"갈라디아서","ephesians":"에베소서","philippians":"빌립보서",
    "colossians":"골로새서",
    "first-thessalonians":"데살로니가전서","second-thessalonians":"데살로니가후서",
    "first-timothy":"디모데전서","second-timothy":"디모데후서",
    "titus":"디도서","philemon":"빌레몬서","hebrews":"히브리서","james":"야고보서",
    "first-peter":"베드로전서","second-peter":"베드로후서",
    "first-john":"요한일서","second-john":"요한이서","third-john":"요한삼서",
    "jude":"유다서","revelation":"요한계시록",
}
def folder_to_book_id(name: str) -> str:
    return re.sub(r"[\s_\-]", "", name).lower()

# ══════════════════════════════════════════════════════════════════
# 1. 한글 자모 분해 + 유사도
# ══════════════════════════════════════════════════════════════════
_CHO = list("ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ")
_JUNG = list("ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ")
_JONG = [""] + list("ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ")

def _decompose(s: str) -> str:
    out = []
    for ch in s:
        c = ord(ch)
        if 0xAC00 <= c <= 0xD7A3:
            i = c - 0xAC00
            out.append(_CHO[i//588] + _JUNG[(i%588)//28] + _JONG[i%28])
        elif ch.strip():
            out.append(ch)
    return "".join(out)

def _norm(s) -> str:
    if s is None: return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s)))

def sim(a: str, b: str) -> float:
    da, db = _decompose(_norm(a)), _decompose(_norm(b))
    if not da or not db: return 0.0
    return (1 - Levenshtein.distance(da, db) / max(len(da), len(db))) * 100

# ══════════════════════════════════════════════════════════════════
# 2. OCR 훅  (여기를 본인 엔진으로 교체하거나, 사이드카 .txt 사용)
# ══════════════════════════════════════════════════════════════════
def ocr_image(path: Path) -> str:
    """이미지 한 장 → 이름 문자열. 기본은 미구현(빈 문자열).
    본인 손글씨 OCR 엔진 호출로 이 본문을 교체하세요.
    예) easyocr:
        import easyocr; global _R
        if '_R' not in globals(): _R = easyocr.Reader(['ko'])
        return ' '.join(_R.readtext(str(path), detail=0))
    """
    return ""   # 사이드카 .txt가 있으면 그쪽이 우선 사용됨

def read_name(img: Path) -> str:
    side = img.with_suffix(".txt")
    if side.exists():
        return side.read_text(encoding="utf-8").strip()
    return ocr_image(img).strip()

# name_raw 정제:  "이름" 라벨 접두 제거 + 모든 공백 제거
_LABEL = re.compile(r"^\s*이\s*름\s*")
def clean_name_raw(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    s = _LABEL.sub("", s)            # 앞머리 '이름' 라벨 제거
    return re.sub(r"\s+", "", s).strip()

def load_jsonl(path: Path):
    """OCR jsonl → [(book_id, book_ko, [(page, raw, source_path) ...]) ...].
    한 줄 예: {"book_id","page","name_raw","engine","source_path","error"}"""
    import json
    books = {}
    for ln in Path(path).read_text(encoding="utf-8").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        r = json.loads(ln)
        raw = "" if r.get("error") else clean_name_raw(r.get("name_raw", ""))
        books.setdefault(r["book_id"], []).append(
            (int(r["page"]), raw, r.get("source_path", "")))
    result = []
    for bid, recs in sorted(books.items()):
        recs.sort(key=lambda x: x[0])          # 페이지 순서로 정렬 (파일이 뒤섞여 있음)
        result.append((bid, BOOK_KO.get(bid), recs))
    return result

# ══════════════════════════════════════════════════════════════════
# 3. 분배표 로더 + 매처
# ══════════════════════════════════════════════════════════════════
def _fmt(v):
    try: return str(int(v))
    except (ValueError, TypeError): return str(v)

class Table:
    def __init__(self, xlsx):
        df = pd.read_excel(xlsx)
        df["_norm"] = df["Name"].map(_norm)
        self.df = df

    def assigned_str(self, name, book_ko):
        r = self.df[(self.df.Name == name) & (self.df.Book == book_ko)]
        if not len(r): return None
        out = []
        for row in r.itertuples():
            rng = _fmt(row.FROM) + (f"~{_fmt(row.TO)}" if pd.notna(row.TO) else "")
            out.append(f"{book_ko} {rng}장")
        return "; ".join(out)

    def match(self, ocr, book_ko):
        pool = self.df[self.df.Book == book_ko]
        book_ok = len(pool) > 0
        if not book_ok:                      # 매핑 안 된 책이면 전체 대상
            pool = self.df
        scored = sorted(((sim(ocr, r.Name), r.Name) for r in pool.itertuples()),
                        key=lambda x: -x[0])
        best_s, best_n = scored[0]
        second_s = scored[1][0] if len(scored) > 1 else 0.0
        # 최고점과 동점인 '서로 다른' 이름이 2명 이상일 때만 동명이인 미해결
        tie_names = sorted({n for s, n in scored if abs(s - best_s) < 1e-6})
        return {
            "name": best_n, "score": round(best_s, 1),
            "margin": round(best_s - second_s, 1),
            "book_matched": book_ok, "ambiguous_ties": tie_names if len(tie_names) > 1 else [],
            "top3": [(n, round(s, 1)) for s, n in scored[:3]],
        }

# ══════════════════════════════════════════════════════════════════
# 4. 페이지 번호 추출 + 폴더 수집
# ══════════════════════════════════════════════════════════════════
def page_num(img: Path, fallback: int) -> int:
    nums = re.findall(r"\d+", img.stem)
    return int(nums[-1]) if nums else fallback

def collect(root: Path):
    """폴더 모드: root 아래 책 폴더별로 jsonl과 동일한 형식으로 반환
       (book_id, book_ko, [(page, raw, source_path) ...])."""
    books = {}
    for img in root.rglob("*"):
        if img.suffix.lower() not in IMG_EXT:
            continue
        rel = img.relative_to(root)
        book_folder = rel.parts[0] if len(rel.parts) > 1 else "_root"
        books.setdefault(book_folder, []).append(img)
    result = []
    for folder, imgs in sorted(books.items()):
        imgs_sorted = sorted(imgs, key=lambda p: (page_num(p, 0), p.name))
        recs = [(page_num(p, i + 1), clean_name_raw(read_name(p)), str(p))
                for i, p in enumerate(imgs_sorted)]
        bid = folder_to_book_id(folder)
        result.append((bid, BOOK_KO.get(bid), recs))
    return result

# ══════════════════════════════════════════════════════════════════
# 5. 메인 파이프라인
# ══════════════════════════════════════════════════════════════════
def build(source, table: Table):
    """source = [(book_id, book_ko, [(page, raw, source_path) ...]) ...]"""
    writers = {}   # writer_id -> {"name":..., "entries":[...]}

    for book_id, book_ko, recs in source:
        # ---- 5a. 페이지별 이름 확보 & 보정 ----
        seq = []   # per page: dict
        for pg, raw, src in recs:
            rec = {"page": pg, "img": Path(src).name if src else "", "raw": raw,
                   "name": None, "score": 0.0, "margin": 0.0,
                   "blank": (raw == ""), "review": [], "ties": []}
            if raw and book_ko:
                m = table.match(raw, book_ko)
                rec.update(name=m["name"], score=m["score"], margin=m["margin"],
                           ties=m["ambiguous_ties"])
                if m["margin"] < MARGIN_REVIEW:
                    rec["review"].append(f"애매(2위와 {m['margin']}차)")   # 검수
                if m["score"] < SCORE_HARDFLOOR:
                    rec["review"].append(f"극저점수({m['score']})")        # 검수
                elif m["score"] < SCORE_NOTE:
                    rec["review"].append(f"저점수({m['score']})")          # 안내(note)
                if m["ambiguous_ties"]:
                    rec["review"].append(f"상위후보동점:{m['ambiguous_ties']}")
                if not m["book_matched"]:
                    rec["review"].append("책매핑없음(전체대상매칭)")
            elif raw and not book_ko:
                rec["review"].append(f"미지의책폴더:{book_id}")
            seq.append(rec)

        # ---- 5b. 빈칸 채우기: 정책상 '앞사람에게 배정' ----
        for i, rec in enumerate(seq):
            if not rec["blank"]:
                continue
            prev = next((seq[j]["name"] for j in range(i-1, -1, -1) if seq[j]["name"]), None)
            nxt  = next((seq[j]["name"] for j in range(i+1, len(seq)) if seq[j]["name"]), None)
            if prev:                                  # 앞사람에게 배정 (기본 규칙)
                rec["name"] = prev
                rec["review"].append("빈칸-앞사람배정")
            elif nxt:                                 # 책 첫머리 빈칸은 앞사람이 없어 뒷사람으로
                rec["name"] = nxt
                rec["review"].append("빈칸-첫머리뒷사람배정")

        # ---- 5b-2. 위치 기반 보정 (같은 책 + 페이지가 실제로 연속일 때만) ----
        named = [r for r in seq if r["name"] is not None]
        runs = []  # 각 run: {"name","recs","start","end"}
        for r in named:
            if runs and runs[-1]["name"] == r["name"]:
                runs[-1]["recs"].append(r); runs[-1]["end"] = r["page"]
            else:
                runs.append({"name": r["name"], "recs": [r],
                             "start": r["page"], "end": r["page"]})
        for t, run in enumerate(runs):
            nm = run["name"]
            Lrun = runs[t-1] if t > 0 else None
            Rrun = runs[t+1] if t < len(runs)-1 else None
            minscore = min((r["score"] for r in run["recs"]), default=100)
            # 페이지가 실제로 맞닿아 있는지 (스캔 공백이면 이웃으로 안 봄)
            L_adj = Lrun and Lrun["end"] + 1 == run["start"]
            R_adj = Rrun and run["end"] + 1 == Rrun["start"]
            if L_adj and R_adj and Lrun["name"] == Rrun["name"] != nm:
                # 샌드위치(양쪽 동일·연속) → 그 사람 run이 갈라진 것 → 흡수 (점수 무관)
                tgt = Lrun["name"]
                for r in run["recs"]:
                    r["review"] = [x for x in r["review"] if not any(
                        k in x for k in ("저점수", "애매", "상위후보동점", "극저점수"))]
                    r["name"] = tgt
                    r["review"].append(f"위치보정-샌드위치(원래:{nm}→{tgt})")
            elif (len(run["recs"]) == 1 and minscore < POS_CASEB_FLOOR
                  and L_adj and (not R_adj or Lrun["name"] != Rrun["name"])):
                # 경계 + 이름 거의 소실(점수<30) + 앞run과 연속 → 앞사람으로
                tgt = Lrun["name"]
                r = run["recs"][0]
                r["review"] = [x for x in r["review"] if not any(
                    k in x for k in ("저점수", "애매", "상위후보동점", "극저점수"))]
                r["name"] = tgt
                r["review"].append(f"위치보정-경계앞사람(원래:{nm}→{tgt},점수{minscore})")

        # ---- 5c. 연속 동일인 → entry 로 묶기 ----
        run = None
        for rec in seq:
            nm = rec["name"]
            if nm is None:
                continue
            if run and run["_name"] == nm:
                run["end_page"] = rec["page"]
                run["_recs"].append(rec)
            else:
                if run: _flush(run, writers, table, book_id, book_ko)
                run = {"_name": nm, "start_page": rec["page"],
                       "end_page": rec["page"], "_recs": [rec]}
        if run: _flush(run, writers, table, book_id, book_ko)

    return list(writers.values())

def _flush(run, writers, table, book_id, book_ko):
    name_id = run["_name"]                 # 분배표 Name = 고유 식별자
    display = re.sub(r"[A-Z]$", "", name_id)
    recs = run["_recs"]
    scores = [r["score"] for r in recs if r["score"]]
    margins = [r["margin"] for r in recs if r["score"]]
    assigned = table.assigned_str(name_id, book_ko) if book_ko else None

    # 검수필요(needs_review) 사유 vs 단순 안내(note) 분리
    # 정책: 동명이인·빈칸배정은 검수 제외(note로만 기록). '저점수'는 안내, '극저점수'는 검수.
    REVIEW_KEYS = ("애매", "극저점수", "책매핑없음", "미지의책", "분배표에없는")
    all_flags = sorted({rr for r in recs for rr in r["review"]})
    review_reasons = [f for f in all_flags if any(k in f for k in REVIEW_KEYS)]
    notes = [f for f in all_flags if f not in review_reasons]
    # 이름이 이 책에 배정된 적 없음 → 매칭 오류 가능성 높음
    if book_ko and assigned is None:
        review_reasons.append(f"분배표에없는배정({name_id}↔{book_ko})")

    entry = {
        "book_id": book_id,
        "page": run["start_page"],
        "start_page": run["start_page"],
        "end_page": run["end_page"],
        "page_count": run["end_page"] - run["start_page"] + 1,
        "assigned": assigned,
        "raw_ocr": [r["raw"] for r in recs],
        "match_score_min": min(scores) if scores else 0.0,
        "match_score_avg": round(sum(scores)/len(scores), 1) if scores else 0.0,
        "match_margin_min": min(margins) if margins else 0.0,
        "needs_review": bool(review_reasons),
        "review_reasons": review_reasons,
        "notes": notes,
    }
    w = writers.setdefault(name_id, {"writer_id": name_id, "name": display, "entries": []})
    w["entries"].append(entry)


# ══════════════════════════════════════════════════════════════════
# 5d. 분리등장 검사 (한 책 = 한 번 연속 등장 규칙)
#   같은 writer가 한 책에서 2개 이상 entry로 갈라져 나오면 = 실존이름 오인식 등의 신호.
#   갈라진 entry를 전부 검수로 올림. (오인식 페이지 양옆 사람이 함께 잡히는 효과)
# ══════════════════════════════════════════════════════════════════
def flag_split_occurrences(writers):
    for w in writers:
        by_book = {}
        for e in w["entries"]:
            by_book.setdefault(e["book_id"], []).append(e)
        for book_id, ents in by_book.items():
            if len(ents) < 2:
                continue
            ents.sort(key=lambda e: e["start_page"])
            spans = ", ".join(f"p{e['start_page']}-{e['end_page']}" for e in ents)
            longest = max(e["page_count"] for e in ents)
            for e in ents:
                tag = f"분리등장({w['writer_id']}@{book_id}:{spans})"
                if e["page_count"] == longest:
                    # 가장 긴 구간(들) = 진짜 등장으로 추정 → 안내만 (단, 최장이 여럿이면 다 검수)
                    n_longest = sum(1 for x in ents if x["page_count"] == longest)
                    if n_longest > 1:
                        e["review_reasons"].append(tag)
                        e["needs_review"] = True
                    else:
                        e["notes"].append(tag + "|본구간추정")
                else:
                    # 짧은/고립 구간 = 오인식 의심 → 검수
                    e["review_reasons"].append(tag + "|고립구간의심")
                    e["needs_review"] = True
    return writers

# ══════════════════════════════════════════════════════════════════
# 6. 검수용 엑셀 출력 (needs_review=true 인 entry만)
# ══════════════════════════════════════════════════════════════════
def write_review_xlsx(writers, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = []
    for w in writers:
        for e in w["entries"]:
            if e["needs_review"]:
                rows.append((w, e))
    # 급한 것부터: 격차 작은 순 → 점수 낮은 순
    rows.sort(key=lambda we: (we[1]["match_margin_min"], we[1]["match_score_min"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "검수필요"

    headers = ["순번", "writer_id", "매칭이름", "book_id", "시작p", "끝p", "장수",
               "배정(분배표)", "원본OCR", "최저점수", "최저격차", "검수사유",
               "★정정이름", "★확인"]
    HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    HFILL = PatternFill("solid", fgColor="4472C4")
    EDIT  = PatternFill("solid", fgColor="FFF2CC")   # 채워넣을 칸(노랑)
    CELLF = Font(name="맑은 고딕", size=10)
    thin  = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = ("검수 안내: 회색 열은 자동 생성값입니다. 노란색 두 열(★)만 채우세요. "
                "'매칭이름'이 맞으면 ★정정이름은 비워두고 ★확인에 O, 틀리면 ★정정이름에 올바른 이름을 적으세요.")
    ws["A1"].font = Font(name="맑은 고딕", size=9, italic=True, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    hr = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HFONT; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for i, (w, e) in enumerate(rows, 1):
        r = hr + i
        vals = [i, w["writer_id"], w["name"], e["book_id"],
                e["start_page"], e["end_page"], e["page_count"],
                e["assigned"] or "", " | ".join(x for x in e["raw_ocr"]),
                e["match_score_min"], e["match_margin_min"],
                "; ".join(e["review_reasons"]), "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELLF; cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (8, 9, 12)))
            if c in (13, 14):                # ★ 편집 칸
                cell.fill = EDIT

    widths = [5, 10, 9, 12, 6, 6, 6, 16, 22, 8, 8, 24, 11, 6]
    for c, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = wd
    ws.freeze_panes = f"A{hr+1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{hr+len(rows)}"

    wb.save(path)
    return len(rows)


# ══════════════════════════════════════════════════════════════════
# 7. 누락 리포트 (분배표엔 있는데 writers.json엔 없는 배정)
#    ① 미수집 책: 그 책 스캔 자체가 없음(책 단위) → 별도 시트로 요약
#    ② 필사누락 의심: 스캔은 있는 책인데 그 사람만 없음 → 실제 조치 대상
# ══════════════════════════════════════════════════════════════════
def write_missing_xlsx(writers, table, source, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 분배표 배정 전체 (Name, Book_ko) → 배정범위 문자열
    assigned = {}
    for r in table.df.itertuples():
        rng = _fmt(r.FROM) + (f"~{_fmt(r.TO)}" if pd.notna(r.TO) else "")
        assigned.setdefault((r.Name, r.Book), f"{r.Book} {rng}장")

    covered = set()          # writers.json이 커버한 (writer_id, book_ko)
    scanned_ko = set()       # 스캔이 실제 존재하는 책(=entry가 하나라도 있는 책)
    for w in writers:
        for e in w["entries"]:
            bko = BOOK_KO.get(e["book_id"])
            covered.add((w["writer_id"], bko))
            if bko:
                scanned_ko.add(bko)

    # 책(book_ko) → 그 책의 OCR 줄 [(page, raw)] 인덱스
    book_lines = {}
    for book_id, book_ko, recs in source:
        if not book_ko:
            continue
        book_lines.setdefault(book_ko, []).extend(
            (pg, raw) for pg, raw, _src in recs if raw)

    def misrecog_candidates(name, book_ko, thresh=45.0, topn=3):
        """그 책 OCR 줄 중 name과 비슷하게 읽히는 것(=오인식 의심) 상위 topn."""
        cands = []
        for pg, raw in book_lines.get(book_ko, []):
            s = sim(raw, name)
            if s >= thresh:
                got = table.match(raw, book_ko)["name"]   # 현재 붙은 이름
                cands.append((s, pg, raw, got))
        cands.sort(key=lambda x: -x[0])
        return cands[:topn]

    missing = [(nm, bk, rng) for (nm, bk), rng in assigned.items()
               if (nm, bk) not in covered]
    individual = sorted([m for m in missing if m[1] in scanned_ko],
                        key=lambda m: (m[1], m[0]))
    wholebook = {}
    for nm, bk, rng in missing:
        if bk not in scanned_ko:
            wholebook.setdefault(bk, []).append(nm)

    # ── 스타일 ──
    HF = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    HFILL = PatternFill("solid", fgColor="C00000")
    CF = Font(name="맑은 고딕", size=10)
    thin = Side(style="thin", color="D9D9D9")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HF; cell.fill = HFILL; cell.border = BD
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # 시트1: 필사누락 의심 (개인 단위, 실제 조치 대상)
    ws1 = wb.active; ws1.title = "필사누락의심"
    ws1["A1"] = ("스캔은 있는 책인데 이 사람 필사본만 없음. "
                 "'오인식후보'에 내용이 있으면 그 페이지가 이 사람인데 다른 이름으로 잘못 읽혔을 가능성, "
                 "비어있으면 진짜 미필사 가능성.")
    ws1["A1"].font = Font(name="맑은 고딕", size=9, italic=True, color="C00000")
    ws1.merge_cells("A1:E1")
    style_header(ws1, ["이름", "책", "배정범위",
                       "오인식후보(그 책 OCR 중 유사)", "비고"], row=2)
    for i, (nm, bk, rng) in enumerate(individual, 1):
        cand = misrecog_candidates(nm, bk)
        cand_str = " / ".join(
            f"p{pg} '{raw}'→현재:{got}(유사{s:.0f})" for s, pg, raw, got in cand
        ) if cand else "(유사 OCR 없음 → 진짜 미필사 가능성)"
        for c, v in enumerate([nm, bk, rng, cand_str, ""], 1):
            cell = ws1.cell(row=2 + i, column=c, value=v)
            cell.font = CF; cell.border = BD
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 4))
    for c, wd in enumerate([12, 12, 16, 46, 18], 1):
        ws1.column_dimensions[get_column_letter(c)].width = wd
    ws1.freeze_panes = "A3"

    # 시트2: 미수집 책 (책 단위 요약)
    ws2 = wb.create_sheet("미수집책")
    ws2["A1"] = "책 스캔/OCR 자체가 없음 → 개인 누락이 아니라 수집 대기"
    ws2["A1"].font = Font(name="맑은 고딕", size=9, italic=True, color="C00000")
    ws2.merge_cells("A1:C1")
    style_header(ws2, ["책", "배정자수", "배정자명단"], row=2)
    for i, (bk, names) in enumerate(sorted(wholebook.items(),
                                    key=lambda x: -len(x[1])), 1):
        for c, v in enumerate([bk, len(names), ", ".join(sorted(names))], 1):
            cell = ws2.cell(row=2 + i, column=c, value=v)
            cell.font = CF; cell.border = BD
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))
    for c, wd in enumerate([14, 9, 80], 1):
        ws2.column_dimensions[get_column_letter(c)].width = wd
    ws2.freeze_panes = "A3"

    wb.save(path)
    return len(individual), sum(len(v) for v in wholebook.values())


# ══════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(
        description="cropped 폴더 또는 OCR jsonl → writers.json")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--jsonl", help="OCR 결과 jsonl (book_id/page/name_raw)")
    src.add_argument("--dir", help="cropped 폴더 (책별 하위폴더, 사이드카 .txt 또는 ocr_image 훅)")
    ap.add_argument("--table", required=True, help="분배표 xlsx 경로")
    ap.add_argument("-o", "--out", default="writers.json")
    ap.add_argument("--review-xlsx", help="검수필요 entry만 담은 엑셀 경로 (예: review.xlsx)")
    ap.add_argument("--missing-xlsx", help="분배표엔 있으나 스캔에 없는 배정 리포트 (예: missing.xlsx)")
    a = ap.parse_args()

    table = Table(a.table)
    source = load_jsonl(Path(a.jsonl)) if a.jsonl else collect(Path(a.dir))
    writers = build(source, table)
    writers = flag_split_occurrences(writers)   # 5d
    Path(a.out).write_text(json.dumps(writers, ensure_ascii=False, indent=2),
                           encoding="utf-8")

    n_rev = sum(1 for w in writers for e in w["entries"] if e["needs_review"])
    n_ent = sum(len(w["entries"]) for w in writers)
    print(f"작성자 {len(writers)}명 · entry {n_ent}개 · 검수필요 {n_rev}개 → {a.out}")

    if a.review_xlsx:
        n = write_review_xlsx(writers, a.review_xlsx)
        print(f"검수 엑셀 {n}행 → {a.review_xlsx}")
    if a.missing_xlsx:
        ind, whole = write_missing_xlsx(writers, table, source, a.missing_xlsx)
        print(f"누락 리포트: 필사누락의심 {ind}명 · 미수집책 배정자 {whole}명 → {a.missing_xlsx}")

if __name__ == "__main__":
    main()