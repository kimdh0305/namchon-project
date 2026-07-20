# -*- coding: utf-8 -*-
"""Build data/writers.json from OCR JSONL records.

This builder supports three layers:
- OCR JSONL input
- optional assignment table input (JSON or CSV)
- optional manual override input (JSON)

It emits the search index used by the Reader and, when requested, an
unmatched report for assignment rows that never produced a writer match.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

from config import ASSIGNMENT_COLUMN_KEYWORDS, ASSIGNMENT_NAME_FUZZY_CUTOFF, BOOK_NAME_TO_ID


def normalize_name(raw):
    s = str(raw or "").strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"^[\"'`“”‘’\[\]\(\)\{\}<>\-_=\+\|:;,.!?~]+", "", s)
    s = re.sub(r"[\"'`“”‘’\[\]\(\)\{\}<>\-_=\+\|:;,.!?~]+$", "", s)
    s = re.sub(r"^(이름|성명)\s*[:：]?\s*", "", s)
    s = re.sub(r"\s*(님|집사|권사|장로|목사)$", "", s)
    s = re.sub(r"[^가-힣]", "", s)
    return s.strip()


def normalize_search_key(raw):
    return re.sub(r"\s+", "", normalize_name(raw)).lower()


def suffix_alpha(idx):
    letters = []
    n = idx
    while True:
        letters.append(chr(ord("A") + (n % 26)))
        n = n // 26 - 1
        if n < 0:
            break
    return "".join(reversed(letters))


def compress_ranges(pages):
    if not pages:
        return []
    pages = sorted(set(int(p) for p in pages))
    ranges = []
    start = prev = pages[0]
    for page in pages[1:]:
        if page == prev + 1:
            prev = page
            continue
        ranges.append((start, prev))
        start = prev = page
    ranges.append((start, prev))
    return ranges


def similarity(left, right):
    return SequenceMatcher(None, normalize_search_key(left), normalize_search_key(right)).ratio()


def parse_int(value):
    if value in (None, ""):
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def normalize_book_id(raw):
    value = str(raw or "").strip()
    if not value:
        return ""
    lowered = value.lower()
    if lowered in BOOK_NAME_TO_ID.values():
        return lowered
    normalized = normalize_name(value)
    return BOOK_NAME_TO_ID.get(normalized, lowered)


def load_json(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path, payload):
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


@dataclass
class AssignmentRow:
    assignment_name: str
    canonical_name: str
    book_id: str
    assign_start: int | None
    assign_end: int | None
    assign_unit: str
    group: str
    raw: dict


def load_records(path):
    rows = []
    with open(path, "r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except Exception:
                continue
            if record.get("error"):
                continue

            book_id = str(record.get("book_id") or "").strip()
            if not book_id:
                continue
            page = parse_int(record.get("page"))
            if page is None:
                continue

            raw_name = normalize_name(record.get("name_raw", ""))
            rows.append(
                {
                    "book_id": book_id,
                    "page": page,
                    "name_raw": raw_name,
                    "name_key": normalize_search_key(raw_name),
                    "source_path": record.get("source_path", ""),
                }
            )
    return rows


def detect_assignment_columns(fieldnames):
    lowered = {name: str(name).strip().lower() for name in fieldnames or []}
    detected = {}
    for target, keywords in ASSIGNMENT_COLUMN_KEYWORDS.items():
        for field, lowered_name in lowered.items():
            if any(keyword.lower() in lowered_name for keyword in keywords):
                detected[target] = field
                break
    return detected


def load_xlsx_rows(path):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        item = {}
        for index, key in enumerate(header):
            if not key:
                continue
            item[key] = row[index] if index < len(row) else None
        if any(value not in (None, "") for value in item.values()):
            data_rows.append(item)
    return data_rows


def load_assignment_rows(path):
    if not path:
        return []

    assignment_path = Path(path)
    suffix = assignment_path.suffix.lower()
    raw_rows = []
    if suffix == ".json":
        payload = load_json(assignment_path)
        raw_rows = payload if isinstance(payload, list) else payload.get("rows", [])
    elif suffix == ".csv":
        with open(assignment_path, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            raw_rows = list(reader)
    elif suffix == ".xlsx":
        raw_rows = load_xlsx_rows(assignment_path)
    else:
        raise ValueError("assignment file must be .json, .csv, or .xlsx")

    if not raw_rows:
        return []

    columns = detect_assignment_columns(raw_rows[0].keys())
    rows = []
    for raw in raw_rows:
        name = normalize_name(raw.get(columns.get("name", ""), ""))
        book_id = normalize_book_id(raw.get(columns.get("book", ""), ""))
        if not name or not book_id:
            continue

        page_value = parse_int(raw.get(columns.get("page", "")))
        start = parse_int(raw.get(columns.get("from", "")))
        end = parse_int(raw.get(columns.get("to", "")))
        chapter = parse_int(raw.get(columns.get("chapter", "")))
        assign_unit = "page"
        if chapter is not None and start is None and end is None and page_value is None:
            start = chapter
            end = chapter
            assign_unit = "chapter"
        elif page_value is not None and start is None and end is None:
            start = page_value
            end = page_value

        rows.append(
            AssignmentRow(
                assignment_name=name,
                canonical_name=name,
                book_id=book_id,
                assign_start=start,
                assign_end=end if end is not None else start,
                assign_unit=assign_unit,
                group=str(raw.get(columns.get("group", ""), "")).strip(),
                raw=dict(raw),
            )
        )
    return rows


def load_overrides(path):
    if not path:
        return {"page_overrides": [], "name_alias_overrides": [], "assignment_overrides": []}
    payload = load_json(path)
    return {
        "page_overrides": payload.get("page_overrides", []),
        "name_alias_overrides": payload.get("name_alias_overrides", []),
        "assignment_overrides": payload.get("assignment_overrides", []),
    }


def apply_assignment_overrides(assignments, overrides):
    override_rows = overrides.get("assignment_overrides", [])
    if not override_rows:
        return assignments

    updated = []
    for assignment in assignments:
        canonical_name = assignment.canonical_name
        for override in override_rows:
            if str(override.get("book_id", "")).strip() != assignment.book_id:
                continue
            if normalize_search_key(override.get("assignment_name", "")) != normalize_search_key(assignment.assignment_name):
                continue
            if parse_int(override.get("assign_start")) != assignment.assign_start:
                continue
            if parse_int(override.get("assign_end")) != assignment.assign_end:
                continue
            canonical_name = normalize_name(override.get("canonical_name", canonical_name)) or canonical_name
            break
        updated.append(
            AssignmentRow(
                assignment_name=assignment.assignment_name,
                canonical_name=canonical_name,
                book_id=assignment.book_id,
                assign_start=assignment.assign_start,
                assign_end=assignment.assign_end,
                assign_unit=assignment.assign_unit,
                group=assignment.group,
                raw=assignment.raw,
            )
        )
    return updated


def build_page_override_map(overrides):
    mapping = {}
    for item in overrides.get("page_overrides", []):
        book_id = str(item.get("book_id", "")).strip()
        page = parse_int(item.get("page"))
        canonical_name = normalize_name(item.get("canonical_name", ""))
        if book_id and page is not None and canonical_name:
            mapping[(book_id, page)] = canonical_name
    return mapping


def build_alias_override_map(overrides):
    mapping = {}
    for item in overrides.get("name_alias_overrides", []):
        raw_name = normalize_search_key(item.get("raw_name", ""))
        canonical_name = normalize_name(item.get("canonical_name", ""))
        scope = item.get("scope") or {}
        book_id = str(scope.get("book_id", "")).strip() or None
        if raw_name and canonical_name:
            mapping[(book_id, raw_name)] = canonical_name
    return mapping


def assignment_candidates_by_book(assignments):
    grouped = defaultdict(list)
    for assignment in assignments:
        grouped[assignment.book_id].append(assignment)
    return grouped


def page_in_assignment(record, assignment):
    if assignment.assign_start is None:
        return True
    if assignment.assign_end is None:
        return record["page"] == assignment.assign_start
    return assignment.assign_start <= record["page"] <= assignment.assign_end


def resolve_canonical_name(record, assignments_by_book, page_overrides, alias_overrides):
    page_override = page_overrides.get((record["book_id"], record["page"]))
    if page_override:
        return page_override, "page_override", []

    alias_override = alias_overrides.get((record["book_id"], record["name_key"])) or alias_overrides.get((None, record["name_key"]))
    if alias_override:
        return alias_override, "alias_override", []

    candidates = assignments_by_book.get(record["book_id"], [])
    if not candidates:
        return record["name_raw"], "ocr_only", []

    if not record["name_raw"]:
        return None, "empty_ocr_name", candidates

    best = None
    best_score = -1.0
    for candidate in candidates:
        score = similarity(record["name_raw"], candidate.assignment_name)
        if score > best_score:
            best = candidate
            best_score = score

    if best and best_score >= ASSIGNMENT_NAME_FUZZY_CUTOFF:
        return best.canonical_name, "assignment_match", candidates
    return None, "no_name_match", candidates


def build_writers(records, assignments, overrides):
    assignments = apply_assignment_overrides(assignments, overrides)
    assignments_by_book = assignment_candidates_by_book(assignments)
    page_overrides = build_page_override_map(overrides)
    alias_overrides = build_alias_override_map(overrides)

    matched_rows = []
    unmatched_rows = []
    matched_assignment_keys = set()

    for record in records:
        canonical_name, reason, candidates = resolve_canonical_name(
            record,
            assignments_by_book,
            page_overrides,
            alias_overrides,
        )
        if canonical_name:
            matched_rows.append(
                {
                    "name": canonical_name,
                    "book_id": record["book_id"],
                    "page": record["page"],
                }
            )
            for candidate in candidates:
                matched_assignment_keys.add(
                    (candidate.book_id, candidate.assignment_name, candidate.assign_start, candidate.assign_end)
                )
            continue

        if candidates:
            indexed_names = {
                row["name"]
                for row in matched_rows
                if row["book_id"] == record["book_id"]
            }
            candidate_writers = []
            seen_candidates = set()
            for candidate in candidates:
                candidate_name = candidate.canonical_name
                if not candidate_name or candidate_name in indexed_names or candidate_name in seen_candidates:
                    continue
                seen_candidates.add(candidate_name)
                candidate_writers.append(candidate_name)
            unmatched_rows.append(
                {
                    "book_id": record["book_id"],
                    "page": record["page"],
                    "raw_name": record["name_raw"],
                    "candidate_writers": candidate_writers,
                    "mismatch_reason": reason,
                    "source_path": record["source_path"],
                }
            )

    by_name = defaultdict(list)
    for row in matched_rows:
        by_name[row["name"]].append(row)

    writers = []
    for index, name in enumerate(sorted(by_name.keys())):
        pages_by_book = defaultdict(list)
        for row in by_name[name]:
            pages_by_book[row["book_id"]].append(row["page"])

        entries = []
        for book_id in sorted(pages_by_book.keys()):
            for start, end in compress_ranges(pages_by_book[book_id]):
                entries.append(
                    {
                        "book_id": book_id,
                        "page": start,
                        "start_page": start,
                        "end_page": end,
                        "line_hint": "center",
                    }
                )

        writers.append(
            {
                "writer_id": f"{name}{suffix_alpha(index)}",
                "name": name,
                "entries": entries,
            }
        )

    unmatched_assignments = []
    for assignment in assignments:
        key = (assignment.book_id, assignment.assignment_name, assignment.assign_start, assignment.assign_end)
        if key in matched_assignment_keys:
            continue
        candidate_pages = [
            record["page"]
            for record in records
            if record["book_id"] == assignment.book_id and page_in_assignment(record, assignment)
        ]
        if not candidate_pages:
            continue
        raw_names_seen = sorted(
            {
                record["name_raw"]
                for record in records
                if record["book_id"] == assignment.book_id and page_in_assignment(record, assignment) and record["name_raw"]
            }
        )
        unmatched_assignments.append(
            {
                "book_id": assignment.book_id,
                "assignment_name": assignment.assignment_name,
                "canonical_name": assignment.canonical_name,
                "assignment_group": assignment.group,
                "assign_start": assignment.assign_start,
                "assign_end": assignment.assign_end,
                "assign_unit": assignment.assign_unit,
                "candidate_pages": candidate_pages,
                "matched_pages": [],
                "raw_names_seen": raw_names_seen,
                "candidate_writers": [assignment.canonical_name],
                "mismatch_reason": "assignment_has_reference_but_no_writer_match",
            }
        )

    return writers, unmatched_rows, unmatched_assignments


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="OCR JSONL path")
    parser.add_argument("--out", default="data/writers.json", help="Output writers.json path")
    parser.add_argument("--assignment", default=None, help="Assignment table path (.json or .csv)")
    parser.add_argument("--overrides", default=None, help="Manual override JSON path")
    parser.add_argument("--report-out", default=None, help="Unmatched report JSON path")
    args = parser.parse_args()

    records = load_records(args.input)
    assignments = load_assignment_rows(args.assignment)
    overrides = load_overrides(args.overrides)
    writers, unmatched_rows, unmatched_assignments = build_writers(records, assignments, overrides)

    write_json(args.out, writers)
    if args.report_out:
        write_json(
            args.report_out,
            {
                "unmatched_records": unmatched_rows,
                "unmatched_assignments": unmatched_assignments,
            },
        )

    print(f"입력 레코드: {len(records)} / 생성 writer: {len(writers)} -> {args.out}")
    if args.report_out:
        print(
            f"미매칭 레코드: {len(unmatched_rows)} / 미매칭 분배표: {len(unmatched_assignments)} -> {args.report_out}"
        )


if __name__ == "__main__":
    main()
